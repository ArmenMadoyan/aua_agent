"""End-to-end RAG evaluation orchestrator.

For each test case: run the RAG pipeline → judge the answer along 5 dimensions
with Gemini → aggregate into a JSON report.
"""

from __future__ import annotations

import json
import logging
import re
import statistics
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from backend.config import ANTHROPIC_MODEL
from backend.tests.eval.gemini_judge import GeminiJudge
from backend.tests.eval.rag_runner import RagResult, RagRunner

logger = logging.getLogger(__name__)

DIMENSIONS = [
    "citation_accuracy",
    "top1_accuracy",
    "hallucination",
    "completeness",
    "relevance",
]
PASS_THRESHOLD = 3  # score >= 3 (out of 5) counts as pass


def _sanitize_for_filename(value: str) -> str:
    """Make a model name safe for a filename (strip path separators, spaces, etc.)."""
    return re.sub(r"[^A-Za-z0-9._-]", "_", value or "unknown_model")


class RAGEvaluator:
    def __init__(
        self,
        test_data_path: str | Path,
        output_dir: str | Path,
        runner: RagRunner | None = None,
        judge: GeminiJudge | None = None,
    ):
        self._test_data_path = Path(test_data_path)
        self._output_dir = Path(output_dir)
        self._runner = runner or RagRunner()
        self._judge = judge or GeminiJudge()

    def _load_queries(self) -> list[dict]:
        with self._test_data_path.open() as f:
            data = json.load(f)
        return data.get("queries", [])

    def _judge_one(
        self, test_case: dict, rag_result: RagResult
    ) -> dict[str, dict[str, Any]]:
        """Run all 5 judges on a single RAG result. Returns dict of dimension → judge output."""
        query = test_case["query"]
        answer = rag_result.answer
        chunks = [
            {"content": c.content, "source": c.source, "score": c.score}
            for c in rag_result.chunks
        ]
        retrieved_sources = [c["source"] for c in chunks]
        top1 = chunks[0] if chunks else {"source": "", "content": ""}

        if rag_result.error or not answer.strip():
            return {
                dim: {
                    "score": None,
                    "explanation": f"skipped (rag_error: {rag_result.error})",
                }
                for dim in DIMENSIONS
            }

        return {
            "citation_accuracy": self._judge.evaluate_citation_accuracy(
                query=query,
                answer=answer,
                retrieved_sources=retrieved_sources,
                expected_policies=test_case.get("expected_relevant_policies", []),
            ),
            "top1_accuracy": self._judge.evaluate_top1_accuracy(
                query=query,
                answer=answer,
                top1_source=top1["source"],
                top1_content=top1["content"],
            ),
            "hallucination": self._judge.evaluate_hallucination(
                query=query, answer=answer, retrieved_chunks=chunks
            ),
            "completeness": self._judge.evaluate_completeness(
                query=query,
                answer=answer,
                expected_topics=test_case.get("expected_topics", []),
            ),
            "relevance": self._judge.evaluate_relevance(query=query, answer=answer),
        }

    def evaluate_all(self, max_queries: int | None = None) -> dict[str, Any]:
        queries = self._load_queries()
        if max_queries is not None:
            queries = queries[:max_queries]

        per_query: list[dict[str, Any]] = []
        for i, test_case in enumerate(queries, start=1):
            qid = test_case.get("id", i)
            logger.info("Evaluating [%d/%d] id=%s", i, len(queries), qid)
            rag_result = self._runner.run(test_case["query"], chat_id=f"eval-{qid}")
            scores = self._judge_one(test_case, rag_result)
            per_query.append(
                {
                    "test_case": test_case,
                    "rag_result": rag_result.to_dict(),
                    "scores": scores,
                }
            )

        return {
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "model": ANTHROPIC_MODEL,
            "judge_model": "gemini-2.5-flash",
            "total_queries": len(per_query),
            "aggregate": self._aggregate(per_query),
            "by_category": self._aggregate_by_category(per_query),
            "per_query": per_query,
        }

    def _scores_for_dimension(self, per_query: list[dict], dimension: str) -> list[int]:
        out: list[int] = []
        for entry in per_query:
            raw = entry["scores"].get(dimension, {})
            s = raw.get("score") if isinstance(raw, dict) else None
            if isinstance(s, (int, float)):
                out.append(int(s))
        return out

    def _dimension_stats(self, scores: list[int]) -> dict[str, Any]:
        if not scores:
            return {
                "mean": None,
                "median": None,
                "min": None,
                "max": None,
                "pass_rate": None,
                "count": 0,
            }
        return {
            "mean": round(statistics.mean(scores), 3),
            "median": statistics.median(scores),
            "min": min(scores),
            "max": max(scores),
            "pass_rate": round(
                sum(1 for s in scores if s >= PASS_THRESHOLD) / len(scores), 3
            ),
            "count": len(scores),
        }

    def _aggregate(self, per_query: list[dict]) -> dict[str, Any]:
        out: dict[str, Any] = {}
        all_means: list[float] = []
        for dim in DIMENSIONS:
            stats = self._dimension_stats(self._scores_for_dimension(per_query, dim))
            out[dim] = stats
            if stats["mean"] is not None:
                all_means.append(stats["mean"])

        latencies = [e["rag_result"]["latency_seconds"] for e in per_query]
        rag_errors = sum(1 for e in per_query if e["rag_result"]["error"])

        out["overall_mean"] = (
            round(statistics.mean(all_means), 3) if all_means else None
        )
        out["avg_latency_seconds"] = (
            round(statistics.mean(latencies), 3) if latencies else None
        )
        out["rag_error_count"] = rag_errors
        return out

    def _aggregate_by_category(self, per_query: list[dict]) -> dict[str, Any]:
        buckets: dict[str, list[dict]] = defaultdict(list)
        for entry in per_query:
            cat = entry["test_case"].get("category", "uncategorized")
            buckets[cat].append(entry)

        return {
            cat: {
                dim: self._dimension_stats(self._scores_for_dimension(entries, dim))
                for dim in DIMENSIONS
            }
            | {"count": len(entries)}
            for cat, entries in buckets.items()
        }

    def _build_filenames(self, report: dict[str, Any]) -> tuple[str, str]:
        """Build (json_name, xlsx_name) with the LLM model name embedded."""
        model_slug = _sanitize_for_filename(report.get("model", ANTHROPIC_MODEL))
        ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        base = f"eval_report_{model_slug}_{ts}"
        return f"{base}.json", f"{base}.xlsx"

    def save_report(
        self,
        report: dict[str, Any],
        json_filename: str | None = None,
        xlsx_filename: str | None = None,
    ) -> dict[str, Path]:
        """Save the report as JSON and Excel. Filenames embed the LLM model name."""
        self._output_dir.mkdir(parents=True, exist_ok=True)
        default_json, default_xlsx = self._build_filenames(report)

        json_path = self._output_dir / (json_filename or default_json)
        with json_path.open("w") as f:
            json.dump(report, f, indent=2, ensure_ascii=False)

        xlsx_path = self._output_dir / (xlsx_filename or default_xlsx)
        self._save_excel(report, xlsx_path)

        return {"json": json_path, "xlsx": xlsx_path}

    def _per_query_headers(self) -> list[str]:
        headers = [
            "id",
            "category",
            "query",
            "agent_used",
            "latency_s",
            "error",
            "retrieved_sources",
            "answer_preview",
        ]
        for dim in DIMENSIONS:
            headers.extend([f"{dim}_score", f"{dim}_explanation"])
        return headers

    def _per_query_row(self, entry: dict[str, Any]) -> list[Any]:
        tc = entry["test_case"]
        rr = entry["rag_result"]
        sc = entry["scores"]
        row: list[Any] = [
            tc.get("id"),
            tc.get("category"),
            tc.get("query"),
            rr.get("agent_used"),
            rr.get("latency_seconds"),
            rr.get("error") or "",
            ", ".join(c.get("source", "") for c in rr.get("chunks", [])),
            (rr.get("answer") or "")[:500],
        ]
        for dim in DIMENSIONS:
            d = sc.get(dim, {}) or {}
            row.append(d.get("score"))
            row.append(d.get("explanation", ""))
        return row

    def _save_excel(self, report: dict[str, Any], path: Path) -> None:
        try:
            from openpyxl import Workbook
        except ImportError:
            logger.warning("openpyxl not installed — skipping Excel report at %s", path)
            return

        wb = Workbook()

        # Sheet 1: Summary
        summary = wb.active
        summary.title = "Summary"
        summary.append(["Field", "Value"])
        summary.append(["Timestamp", report.get("timestamp")])
        summary.append(["LLM model", report.get("model")])
        summary.append(["Judge model", report.get("judge_model")])
        summary.append(["Total queries", report.get("total_queries")])
        summary.append(
            ["RAG errors", report.get("aggregate", {}).get("rag_error_count")]
        )
        summary.append(
            ["Avg latency (s)", report.get("aggregate", {}).get("avg_latency_seconds")]
        )
        summary.append(
            ["Overall mean (0-5)", report.get("aggregate", {}).get("overall_mean")]
        )
        summary.append([])
        summary.append(
            ["Dimension", "Mean", "Median", "Min", "Max", "Pass rate (>=3)", "N"]
        )
        for dim in DIMENSIONS:
            s = report["aggregate"].get(dim, {})
            summary.append(
                [
                    dim,
                    s.get("mean"),
                    s.get("median"),
                    s.get("min"),
                    s.get("max"),
                    s.get("pass_rate"),
                    s.get("count"),
                ]
            )

        # Sheet 2: Per Query
        per_q = wb.create_sheet("Per Query")
        per_q.append(self._per_query_headers())
        for entry in report.get("per_query", []):
            per_q.append(self._per_query_row(entry))

        # Sheet 3: By Category
        cat = wb.create_sheet("By Category")
        cat.append(
            ["category", "count"]
            + [f"{dim}_mean" for dim in DIMENSIONS]
            + [f"{dim}_pass_rate" for dim in DIMENSIONS]
        )
        for category, stats in (report.get("by_category") or {}).items():
            row = [category, stats.get("count")]
            row.extend(stats.get(dim, {}).get("mean") for dim in DIMENSIONS)
            row.extend(stats.get(dim, {}).get("pass_rate") for dim in DIMENSIONS)
            cat.append(row)

        wb.save(path)

    @staticmethod
    def print_summary(report: dict[str, Any]) -> None:
        agg = report["aggregate"]
        print("\n" + "=" * 60)
        print("AUA RAG Evaluation Summary")
        print("=" * 60)
        print(f"Total queries: {report['total_queries']}")
        print(f"RAG errors:    {agg.get('rag_error_count', 0)}")
        print(f"Avg latency:   {agg.get('avg_latency_seconds')} s")
        print("")
        print(f"{'dimension':<22}{'mean':>8}{'median':>9}{'pass_rate':>12}{'n':>6}")
        for dim in DIMENSIONS:
            s = agg.get(dim, {})
            print(
                f"{dim:<22}"
                f"{str(s.get('mean')):>8}"
                f"{str(s.get('median')):>9}"
                f"{str(s.get('pass_rate')):>12}"
                f"{str(s.get('count')):>6}"
            )
        print("")
        print(f"Overall mean:  {agg.get('overall_mean')}/5.0")
        print("=" * 60 + "\n")
